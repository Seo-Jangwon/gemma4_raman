# -*- coding: utf-8 -*-
"""라만_벤치마크.xlsx '문항' 시트 단일 리더.

[왜 따로 두는가 — 2026-08-02]
grade.py / run_bench.py / make_proc_gt.py 세 곳이 각자 `ws.cell(r, 11)` 같은 **열 번호**로
같은 시트를 읽고 있었다. 시트를 손보자(문항_통합 16열 → 문항 9열) 세 곳이 동시에 깨졌고,
그중 grade.py 는 예외를 삼켜 "절차 채점을 건너뜁니다" 한 줄만 남기고 **1차 점수를 조용히
0으로 만들었다**. 채점기가 조용히 반쪽이 되는 것이 가장 나쁜 고장이라, 열은 **이름으로**
찾고 못 찾으면 **죽는다**.

바깥에서 쓰는 건 load_tasks() 하나. 반환 dict 의 키는 열 이름이 아니라 역할 이름이다
(시트 헤더 문구가 또 바뀌어도 소비자 코드는 안 바뀐다).
"""
from __future__ import annotations

import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent
XLSX = ROOT.parent / "라만_벤치마크.xlsx"
SHEET = "문항"

# 역할 → 헤더에 반드시 들어 있어야 하는 조각. 부분일치(공백·괄호 표기 흔들림 흡수).
_COLS = {
    "task_id":  ["문제번호"],
    "prompt":   ["Task"],
    "inputs":   ["입력 파일", "입력파일"],
    "first":    ["1차"],
    "gt":       ["2차"],
    "grading":  ["채점방식", "채점 방식"],
    "manual":   ["사람 개입", "사람개입"],
    "axis":     ["역량축"],
    "score":    ["배점"],
}
# 없어도 되는 열(나중에 추가되는 열은 여기에). 없으면 값이 "" 로 들어간다.
_OPTIONAL = {"hypo_prompt", "hypo_gt"}
_COLS_OPT = {
    "hypo_prompt": ["가정형 Task"],
    "hypo_gt":     ["가정형 GT"],
}

_TASK_RE = re.compile(r"^[TN]\d+$")

# 번호는 있는데 Task 본문이 빈 행. load_tasks() 가 채운다 — 실행기가 경고를 낼 수 있게.
BLANK: list = []


def _header_map(ws) -> dict:
    header = {c: str(ws.cell(1, c).value or "").strip()
              for c in range(1, ws.max_column + 1)}
    out, missing = {}, []
    for role, needles in {**_COLS, **_COLS_OPT}.items():
        hit = next((c for c, h in header.items()
                    if h and any(n in h for n in needles)), None)
        if hit is None:
            if role not in _OPTIONAL:
                missing.append(f"{role}({'/'.join(needles)})")
            continue
        out[role] = hit
    if missing:
        raise KeyError(
            f"'{SHEET}' 시트에서 열을 찾지 못했습니다: {', '.join(missing)}\n"
            f"  현재 헤더: {[h for h in header.values() if h]}\n"
            f"  열을 지우거나 이름을 바꿨다면 benchmark/tasksheet.py 의 _COLS 를 맞추세요."
        )
    return out


def load_tasks(xlsx: Path | None = None) -> dict:
    """{task_id: {prompt, inputs, first, gt, grading, manual, axis, score, hypo_*}}"""
    import openpyxl
    path = Path(xlsx or XLSX)
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    col = _header_map(ws)
    out = {}
    BLANK.clear()
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(r, col["task_id"]).value
        if not isinstance(tid, str) or not _TASK_RE.match(tid.strip()):
            continue
        rec = {}
        for role, c in col.items():
            v = ws.cell(r, c).value
            rec[role] = v if role == "score" else str(v or "").strip()
        for role in _OPTIONAL:
            rec.setdefault(role, "")
        rec["score"] = float(rec["score"] or 2)
        # '불필요' 가 '필요' 를 포함한다 — 부분일치로 검사하면 전 문항이 사람 개입이 된다.
        rec["needs_manual"] = rec["manual"].startswith("필요")
        if not rec["prompt"]:
            # 번호만 있고 본문이 빈 행. 그대로 실행하면 빈 프롬프트를 모델에 던지게 된다.
            BLANK.append(tid.strip())
            continue
        out[tid.strip()] = rec
    if not out:
        raise ValueError(f"'{SHEET}' 시트에서 문항을 하나도 읽지 못했습니다: {path}")
    return out


def tool_names() -> set:
    """실존하는 도구 이름 집합. 산문에서 도구를 뽑을 때 오탐을 거른다."""
    import sys
    sys.path.insert(0, str(ROOT.parent))
    from backend.hw_tools.raman_tool_schemas import RAMAN_TOOLS
    from backend.agents.file_tools import FILE_TOOLS
    return ({t["function"]["name"] for t in RAMAN_TOOLS + FILE_TOOLS}
            | {"search_knowledge_base"})


_WORD = re.compile(r"\b([a-z_]{4,})\b")


def tools_in(text: str, valid: set) -> list:
    return sorted({t for t in _WORD.findall(text or "") if t in valid})


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    t = load_tasks()
    man = [k for k, v in t.items() if v["needs_manual"]]
    print(f"문항 {len(t)}개 / 총 배점 {sum(v['score'] for v in t.values()):.0f}점")
    print(f"사람 개입 {len(man)}개: {', '.join(sorted(man))}")
    if BLANK:
        print(f"[warn] 본문이 빈 행 {len(BLANK)}개(실행 제외): {', '.join(BLANK)}")
