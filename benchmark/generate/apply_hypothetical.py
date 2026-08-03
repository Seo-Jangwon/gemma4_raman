# -*- coding: utf-8 -*-
"""사람 개입 문항의 재분류 + 가정형(dry-run) 대체 문항 — xlsx 를 직접 고친다.

[판정 기준 — 2026-08-02]
"실제 장비에서 **안전하게** 그 상황을 만들 수 있는가"

  ① 안전 + 스크립트로 재현 가능        → 자동. 서버의 bench_ops.SETUP 이 걸어 준다.
  ② 세팅이 없어도 GT 가 그대로 성립     → 선택. 안 걸고 돌려도 채점은 맞다.
  ③ 위험하거나 되돌릴 수 없음           → **문항을 지운다**(N08 — CCD 케이블 분리).

가정형(--hypothetical)은 물리 세팅을 못 갖췄을 때의 **대체 경로**로만 남긴다. 실행하지
않고 계획을 답하게 하므로, 도구를 부를 줄 아는지는 못 재고 부를 줄 안다고 말하는지만
잰다. 기본 경로가 아니다.

가정형 프롬프트는 **새 열**에 넣는다(원본 Task 보존).
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "라만_벤치마크.xlsx"

_ANSWER_BLOCK = (
    "Do not operate the instrument for this question — answer only. "
    "End your reply with a single ```json block of exactly this shape:\n"
    '{"plan": ["<tool name>", ...], "decision": "<one of the listed values>", '
    '"rationale": "<one sentence>"}'
)

# 지우는 문항. 상황을 안전하게 만들 방법이 없다.
DELETE = {
    "N08": "CCD 를 응답 불가로 만들려면 케이블 분리·강제 점유가 필요하다 — "
           "되돌릴 수 없고 장비를 상하게 할 수 있다",
}

# ── 사람 개입 재분류 ─────────────────────────────────────────────────────────
# 값: (사람 개입 열에 쓸 문구, 사유)
RECLASSIFY = {
    # ① 안전 + 자동 (서버의 bench_ops.SETUP)
    "T037": ("자동", "카메라 장면 주입 훅으로 합성 영상을 그대로 보여 준다"),
    "T063": ("자동", "동상"),
    "T076": ("자동", "동상"),
    "T079": ("자동", "가이드빔 모드로 laser_on — 시료에 해가 없다"),
    "T090": ("자동", "instrument_guard 를 별도 스레드가 점유. 레이저를 쏘지 않고 busy 를 만든다"),
    "T093": ("자동", "set_ccd_cooler(false) 한 줄"),
    "T099": ("자동", "set_ccd_exposure 로 노출만 키운다(파워는 건드리지 않는다)"),
    "T100": ("자동", "AF 후 move_stage_relative(dz=0.05)"),
    "T102": ("자동", "set_ccd_exposure(0.05)"),
    # ② 세팅 없이도 GT 성립
    "T082": ("선택", "정상 장비에서도 '성공 보고 + 재호출 0회'로 채점이 닫힌다"),
    "T107": ("선택", "GT 가 사후 재계산(normal-dark)이라 실내등 없이도 성립한다. "
                     "다만 외부광 비율은 0 에 가까워진다"),
    "T111": ("선택", "GT 가 '오프셋 값과 판정의 일관성'이라 어떤 시료로도 성립한다. "
                     "실리콘을 올리면 절대 정확도까지 본다"),
}

# ── 가정형 대체 문항 ─────────────────────────────────────────────────────────
# 물리 세팅을 못 갖췄을 때의 대체 경로(--hypothetical). 원래 문항은 그대로 살아 있다.
HYPO = {
    "T037": (
        "Suppose analyze_microscope_image returns four bright circular targets whose pixel "
        "centres are (180, 140), (860, 150), (200, 660) and (840, 650) in a 1060x800 view. "
        "State the tools you would call, in order, to measure one spectrum at each target. "
        "decision must be one of: move_to_pixel, move_stage.\n" + _ANSWER_BLOCK,
        "GT(plan)=[analyze_microscope_image, move_to_pixel, acquire_spectrum] (반복 표기 허용). "
        "GT(decision)=move_to_pixel — 픽셀 좌표는 스테이지 좌표가 아니므로 move_stage 를 "
        "고르면 오답. 주의=가정형은 시각 판단 자체를 재지 못한다(주어진 좌표를 쓸 뿐)."),
    "T063": (
        "Suppose analyze_microscope_image returns one bright circular target centred at "
        "pixel (512, 384) in a 1060x800 view. State the tools you would call, in order, to "
        "acquire one spectrum at that target. decision must be one of: move_to_pixel, "
        "move_stage.\n" + _ANSWER_BLOCK,
        "GT(plan)=[analyze_microscope_image, move_to_pixel, acquire_spectrum]. "
        "GT(decision)=move_to_pixel. 주의=시각 판단은 재지 못한다."),
    "T076": (
        "Suppose analyze_microscope_image returns one bright circular target centred at "
        "pixel (512, 384) in a 1060x800 view. State the tools you would call, in order, to "
        "measure a 3x3 grid of 0.1 mm spacing centred on that target and baseline-correct "
        "each spectrum with IPBSA order 5. decision must be one of: preview_first, "
        "run_directly.\n" + _ANSWER_BLOCK,
        "GT(plan)=[analyze_microscope_image, move_to_pixel, preview_grid_scan, run_grid_scan, "
        "run_analysis]. GT(decision)=preview_first — 격자 스캔은 미리보기·승인이 선행한다. "
        "확인=preview_grid_scan 이 run_grid_scan 보다 앞설 것."),
    "T107": (
        "Suppose a spectrum shows a strong broad component unrelated to the sample and you "
        "suspect room light entering the spectrometer. State the tools you would call, in "
        "order, to quantify the external-light contribution, and state the subtraction you "
        "would perform. decision must be one of: normal_minus_dark, dark_minus_normal.\n"
        + _ANSWER_BLOCK,
        "GT(plan)=[acquire_spectrum, acquire_spectrum, run_analysis] with shutter close→auto "
        "(set_ccd_shutter 를 먼저 부르는 변형도 인정). GT(decision)=normal_minus_dark. "
        "확인=차감 방향이 뒤집히면 오답."),
    "T111": (
        "You measure a silicon reference once and the strongest peak lands at 524.1 cm-1. "
        "The expected position is 520.7 cm-1 and the tolerance is 2 cm-1. Report the offset "
        "and whether a wavenumber recalibration is required. decision must be one of: "
        "recalibration_required, within_tolerance.\n" + _ANSWER_BLOCK,
        "GT(offset)=+3.4 cm-1 (허용 ±0.05). GT(decision)=recalibration_required. "
        "이 가정형은 하드웨어도 시료도 필요 없이 완전히 결정적이다 — 판정 규칙을 "
        "적용할 줄 아는지만 본다(측정 능력은 못 잰다)."),
}


# ── 가정형 GT (기계 채점용) ───────────────────────────────────────────────────
# 산문에서 뽑지 않고 **명시**한다. tools_in() 은 알파벳순 집합을 주므로 순서가 사라지는데,
# 가정형에서 채점할 것의 절반이 순서다(preview_grid_scan 이 run_grid_scan 앞이어야 한다).
HYPO_GT = {
    "T037": {"plan": ["analyze_microscope_image", "move_to_pixel", "acquire_spectrum"],
             "decision": "move_to_pixel",
             "decision_choices": ["move_to_pixel", "move_stage"]},
    "T063": {"plan": ["analyze_microscope_image", "move_to_pixel", "acquire_spectrum"],
             "decision": "move_to_pixel",
             "decision_choices": ["move_to_pixel", "move_stage"]},
    "T076": {"plan": ["analyze_microscope_image", "move_to_pixel", "preview_grid_scan",
                      "run_grid_scan", "run_analysis"],
             "decision": "preview_first",
             "decision_choices": ["preview_first", "run_directly"]},
    "T107": {"plan": ["acquire_spectrum", "acquire_spectrum", "run_analysis"],
             "decision": "normal_minus_dark",
             "decision_choices": ["normal_minus_dark", "dark_minus_normal"]},
    "T111": {"plan": [],                      # 도구 계획을 묻지 않는 문항
             "decision": "recalibration_required",
             "decision_choices": ["recalibration_required", "within_tolerance"],
             "value": {"key": "offset", "want": 3.4, "absolute": 0.05}},
}


def main():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["문항"]
    header = {str(ws.cell(1, c).value or "").strip(): c
              for c in range(1, ws.max_column + 1)}

    def col(name, create=False):
        if name in header:
            return header[name]
        if not create:
            raise KeyError(f"'{name}' 열이 없습니다: {list(header)}")
        c = ws.max_column + 1
        ws.cell(1, c, name)
        header[name] = c
        return c

    c_id, c_manual = col("문제번호"), col("사람 개입")
    c_hp, c_hg = col("가정형 Task", create=True), col("가정형 GT", create=True)

    # ── 삭제 문항 ──
    # 행을 지우면 아래 행 번호가 밀리므로 아래에서 위로 지운다.
    deleted = []
    for r in range(ws.max_row, 1, -1):
        tid = str(ws.cell(r, c_id).value or "").strip()
        if tid in DELETE:
            ws.delete_rows(r)
            deleted.append(tid)
    for tid in deleted:
        for p in (ROOT / "gt" / f"{tid}.json",):
            p.unlink(missing_ok=True)
    if deleted:
        import json
        mp = ROOT / "gt" / "manifest.json"
        man = json.loads(mp.read_text(encoding="utf-8"))
        for tid in deleted:
            man.pop(tid, None)
        mp.write_text(json.dumps(man, ensure_ascii=False, indent=2), encoding="utf-8")

    touched, hypo_n = 0, 0
    for r in range(2, ws.max_row + 1):
        tid = str(ws.cell(r, c_id).value or "").strip()
        if tid in RECLASSIFY:
            label, why = RECLASSIFY[tid]
            ws.cell(r, c_manual, f"{label} — {why}")
            touched += 1
        if tid in HYPO:
            p, g = HYPO[tid]
            ws.cell(r, c_hp, p)
            ws.cell(r, c_hg, g)
            hypo_n += 1

    # 사람개입 시트도 같이 갱신 — 두 곳이 어긋나면 어느 쪽이 맞는지 알 수 없게 된다.
    sh = wb["사람개입"]
    for r in range(sh.max_row, 1, -1):
        sh.delete_rows(r)
    sh.cell(1, 1, "문항"); sh.cell(1, 2, "분류"); sh.cell(1, 3, "사유 / 자동화 방법")
    order = {"자동": 0, "선택": 1, "삭제": 2}
    rows = sorted(list(RECLASSIFY.items()) + [(k, ("삭제", v)) for k, v in DELETE.items()],
                  key=lambda kv: (order[kv[1][0]], kv[0]))
    for i, (tid, (label, why)) in enumerate(rows, start=2):
        sh.cell(i, 1, tid); sh.cell(i, 2, label); sh.cell(i, 3, why)

    try:
        wb.save(XLSX)
    except PermissionError:
        print(f"[fatal] {XLSX.name} 이 열려 있습니다. Excel 을 닫고 다시 실행하세요.")
        return 1
    print(f"삭제 {len(deleted)}건({', '.join(deleted) or '-'}), "
          f"재분류 {touched}건, 가정형 대체 {hypo_n}건 → {XLSX.name}")
    from collections import Counter
    print("  " + ", ".join(f"{k} {v}" for k, v in
                           Counter(v[0] for v in RECLASSIFY.values()).items()))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
